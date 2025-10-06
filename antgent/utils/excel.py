import io
import logging
from typing import Any

import openpyxl  # type: ignore
from fastapi import HTTPException, UploadFile
from openpyxl import Workbook  # type: ignore
from pydantic import BaseModel, ValidationError

logger = logging.getLogger(__name__)


def _str_to_bool(value: Any) -> Any:
    """Converts string 'TRUE'/'FALSE' to boolean, case-insensitive."""
    if isinstance(value, str):
        val_upper = value.upper()
        if val_upper == "TRUE":
            return True
        if val_upper == "FALSE":
            return False
    return value


def parse_excel_to_models[T: BaseModel](file: UploadFile, model_cls: type[T]) -> list[T]:
    """Parses an XLSX file into a list of Pydantic model objects."""
    if not file.filename or not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="File must be an .xlsx file.")
    try:
        # load_workbook can take a file-like object directly.
        # file.file is a SpooledTemporaryFile which is a file-like object
        workbook = openpyxl.load_workbook(file.file)
        sheet = workbook.active
        if not sheet:
            raise ValueError("The workbook is empty or the active sheet could not be found.")

        # Read headers from the first row
        headers = [cell.value for cell in sheet[1]]
        if not all(headers):
            raise ValueError("The header row contains empty cells.")

        raw_table = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            # Skip empty rows
            if not any(row):
                continue
            row_data = dict(zip(headers, row, strict=False))

            # Sanitize boolean-like strings
            for key, value in row_data.items():
                row_data[key] = _str_to_bool(value)

            raw_table.append(row_data)

        if not raw_table:
            return []

        return [model_cls.model_validate(item) for item in raw_table]
    except ValidationError as e:
        logger.error("Pydantic validation failed during Excel parsing: %s", e, exc_info=True)
        raise HTTPException(status_code=400, detail=f"Excel data does not match expected schema: {e}") from e
    except Exception as e:
        logger.error("Failed to parse Excel file: %s", e, exc_info=True)
        raise HTTPException(status_code=400, detail=f"Failed to parse Excel file: {e}") from e


def list_dict_to_xlsx_bytes(data: list[dict[str, Any]]) -> bytes:
    """
    Converts a list of dictionaries to an XLSX file as bytes.

    Args:
        data: A list of dictionaries. Each dictionary represents a row.
              It's assumed that all dictionaries have the same keys.

    Returns:
        The XLSX file content as bytes.
        Returns empty bytes if the input list is empty.
    """
    if not data:
        return b""

    workbook = Workbook()
    sheet = workbook.active

    # Write headers
    headers = list(data[0].keys())
    for col_num, header_title in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = header_title

    # Write data rows
    for row_num, row_data in enumerate(data, 2):
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=row_num, column=col_num)
            cell.value = row_data.get(header)

    # Save to a bytes buffer
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
